import cv2 as cv
import mediapipe as mp
import numpy as np
import time
import os
import threading
import openpyxl
import datetime

class SquatTracker:
    def __init__(self):
        self.mp_drawing = mp.solutions.drawing_utils
        self.mp_pose = mp.solutions.pose
        self.capture = None
        self.stage = None
        self.counter = 0
        self.rep_speeds = [] 
        self.average_speed = 0
        self.current_rep_time = 0  
        self.last_rep_time = 0 
        self.waitKey_delay = 1
        self.width = 0
        self.height = 0

    def calculate_angle(self, start, mid, end):
        start = np.array(start)
        mid = np.array(mid)
        end = np.array(end)

        radians = np.arctan2(end[1] - mid[1], end[0] - mid[0]) - np.arctan2(start[1] - mid[1], start[0] - mid[0])
        angle = np.abs(radians * 180.0 / np.pi)

        if angle > 180.0:
            angle = 360 - angle

        return angle

    def write_to_excel(self):
        # Get the directory path of the main program
        program_dir = os.path.dirname(os.path.abspath(__name__))

        # Construct the full path for the Excel file in the same directory as the program
        excel_file_path = os.path.join(program_dir, 'squat_data.xlsx')

        try:
            # Check if the Excel file already exists; if so, load it
            if os.path.exists(excel_file_path):
                wb = openpyxl.load_workbook(excel_file_path)
                sheet = wb.active
                print("Excel file exists")
            else:
                print("Excel file does not exist, creating new workbook")
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet['A1'] = 'Date'
                sheet['B1'] = 'Weight Loaded (kg/lb)'
                sheet['C1'] = 'Rep Speed (sec/rep)' 
                sheet['D1'] = 'Avg. Speed (sec/rep)'    

            next_row = sheet.max_row + 1

            sheet.cell(row=next_row, column=1).value = self.date
            sheet.cell(row=next_row, column=2).value = self.weight
            sheet.cell(row=next_row, column=4).value = self.average_speed
            for i, rep_speed in enumerate(self.rep_speeds):
                sheet.cell(row=next_row + i, column=3).value = rep_speed

            wb.save(excel_file_path)
            print(f"Data successfully saved to {excel_file_path}")
        except Exception as e:
            print(f"Error while saving data: {str(e)}")
    
    def get_date(self):
        valid_date = False
        while not valid_date:
            self.date = input("Enter the date (YYYY-MM-DD): ")
            try:
                datetime.datetime.strptime(self.date, '%Y-%m-%d')
                valid_date = True
            except ValueError:
                print("Invalid date format. Please enter a valid date in YYYY-MM-DD format.")

    def get_weight_loaded(self):
        valid_weight = False
        while not valid_weight:
            self.weight = input("Enter the weight loaded on the bar (kg/lb): ")
            try:
                # Assuming the user provides weight in a format like "100 kg" or "225 lb"
                weight, unit = self.weight.split()
                weight = float(weight)
                if unit.lower() in ('kg', 'lb'):
                    valid_weight = True
                else:
                    print("Invalid weight unit. Please enter 'kg' or 'lb'.")
            except ValueError:
                print("Invalid input. Please enter the weight in the format '100 kg' or '225 lb'.")

    def initialize_video_capture(self):
        video_extensions = [".mp4", ".mkv", ".avi", ".mov", ".wmv", ".flv", ".webm", ".m4v", ".mpeg", ".mpg",
                   ".3gp", ".3g2", ".m2v", ".ts", ".ogg", ".ogv", ".qt", ".rm", ".rmvb", ".vob"]
        while True:
            video = input("Enter directory path to video: ")
            if os.path.exists(video):
                _, file_extension = os.path.splitext(video)
                if file_extension.lower() in video_extensions:
                    break
                else:
                    print("Invalid video file format. Please enter a valid video file.")
            else:
                print("File not found. Please enter a valid file path.")
        self.capture = cv.VideoCapture(video)
    
    def start_tracking(self):
        self.initialize_video_capture()
        self.get_date()
        self.get_weight_loaded() 
        self.width = int(self.capture.get(3))
        self.height = int(self.capture.get(4))

        counter = 1
        base_filename = 'squat_analyzed'
        file_extension = '.mp4'
        output_filename = f"{base_filename}_{counter}{file_extension}"
        while os.path.exists(output_filename):
            counter += 1
            output_filename = f"{base_filename}_{counter}{file_extension}"
            
        fourcc = cv.VideoWriter_fourcc(*'mp4v')
        output_video = cv.VideoWriter(output_filename, fourcc, 30.0, (self.width, self.height))

        thread1 = threading.Thread(target=self.process_image, args=(output_video,))
        thread1.start()

    def process_image(self, output_video):
        with self.mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
            while self.capture.isOpened():
                ret, frame = self.capture.read()

                if not ret:
                    break  

                image = cv.cvtColor(frame, cv.COLOR_BGR2RGB)
                image.flags.writeable = False
                results = pose.process(image)
                image.flags.writeable = True        
                image = cv.cvtColor(image, cv.COLOR_RGB2BGR)

                try:
                    landmarks = results.pose_landmarks.landmark
                    
                    left_shoulder = landmarks[self.mp_pose.PoseLandmark.LEFT_SHOULDER.value]
                    left_hip = landmarks[self.mp_pose.PoseLandmark.LEFT_HIP.value]
                    left_knee = landmarks[self.mp_pose.PoseLandmark.LEFT_KNEE.value]

                    angle = self.calculate_angle(
                        [left_shoulder.x, left_shoulder.y],
                        [left_hip.x, left_hip.y],
                        [left_knee.x, left_knee.y]
                    )
                    
                    if angle < 30 and self.stage != "bottom":
                        self.stage = "bottom"                       
                    if 30 < angle < 130 and self.stage == "bottom":
                        self.stage = "ongoing"
                        self.current_rep_time = time.time()
                    if angle > 130 and self.stage == "ongoing":
                        self.stage = "lockout"
                        self.counter += 1
                        self.last_rep_time = time.time() 
                        rep_time = (self.last_rep_time - self.current_rep_time)
                        self.rep_speeds.append(round(rep_time, 2))

                except: 
                    pass

                cv.rectangle(image, (0, 0), (int(self.width), 60), (0,0,0), -1)
                cv.putText(image, "REPS", (5, 25), cv.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 1, cv.LINE_AA)
                cv.putText(image, str(self.counter), (5, 50), cv.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1, cv.LINE_AA)

                cv.putText(image, 'STAGE', (60, 25), cv.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 1, cv.LINE_AA)
                cv.putText(image, self.stage, (60, 50), cv.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 1, cv.LINE_AA)
                
                cv.putText(image, 'AVG SPEED', (170, 25), cv.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 1, cv.LINE_AA)
                if len(self.rep_speeds) > 0:
                    self.average_speed = sum(self.rep_speeds) / len(self.rep_speeds)            
                    cv.putText(image, f'{self.average_speed:.2f} sec/rep', (170, 50), cv.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 1, cv.LINE_AA)

                output_video.write(image)

                cv.imshow("Live Feed", image)
                if cv.waitKey(self.waitKey_delay) & 0xFF == ord('x'):
                    break

        self.write_to_excel()
        output_video.release()
        self.capture.release()
        cv.destroyAllWindows()

if __name__ == '__main__':  
    tracker = SquatTracker()
    tracker.start_tracking()